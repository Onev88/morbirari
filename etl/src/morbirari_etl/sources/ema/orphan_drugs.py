"""Designaciones de medicamento huérfano de la EMA.

Qué es esto exactamente, porque se malinterpreta con facilidad: una designación
huérfana **no** significa que el fármaco esté aprobado, ni disponible, ni que
funcione. Significa que el regulador ha aceptado que se está desarrollando para una
enfermedad rara, lo que da incentivos al laboratorio. Muchas designaciones nunca
llegan a fármaco. La interfaz tiene que decirlo o el dato miente.

Licencia: la EMA publica estas tablas como datos abiertos y se actualizan a diario.

El problema real: **la EMA no publica códigos ORPHA**. La enfermedad viene como texto
libre en «Intended use» ("Treatment of Wilson's disease"). El emparejamiento con
nuestra nomenclatura es, por tanto, una inferencia nuestra y frágil; se guarda con su
método y su confianza, y la interfaz lo presenta como tal.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from typing import Iterator

from morbirari_etl.sources.base import (
    RawArtifact,
    download,
    raw_path,
    remote_fingerprint,
    sha256_file,
    version_from_fingerprint,
)

SOURCE_NAME = "ema-orphan"

URL = (
    "https://www.ema.europa.eu/en/documents/report/"
    "medicines-output-orphan_designations-report_en.xlsx"
)

ATTRIBUTION = "European Medicines Agency (EMA), orphan designations. https://www.ema.europa.eu"

AGENCY = "EMA"


@dataclass(frozen=True)
class StagedDrug:
    designation_number: str
    medicine_name: str | None
    active_substance: str | None
    intended_use: str | None
    status: str | None
    designation_date: str | None
    url: str | None


def fetch(force: bool = False) -> RawArtifact:
    etag, last_modified = remote_fingerprint(URL)
    version = version_from_fingerprint(etag, last_modified)
    dest = raw_path(SOURCE_NAME, version, "ema_orphan_designations.xlsx")
    if not dest.exists() or force:
        download(URL, dest)
    return RawArtifact(
        source=SOURCE_NAME,
        path=dest,
        sha256=sha256_file(dest),
        source_url=URL,
        etag=etag,
    )


def parse(artifact: RawArtifact) -> Iterator[StagedDrug]:
    """Lee el XLSX de la EMA.

    La cabecera no está en la primera fila: el fichero lleva 8 filas de metadatos
    antes. Se localiza buscando la fila que contiene «Medicine name» en vez de fijar
    un número, que se rompería en cuanto la EMA añada una línea.
    """
    import openpyxl

    wb = openpyxl.load_workbook(artifact.path, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    headers: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = [str(c).strip() if c else "" for c in row]
        if "Medicine name" in cells:
            header_row = cells
            headers = cells
            break

    if not header_row:
        raise ValueError("No se encuentra la cabecera del XLSX de la EMA")

    def col(name: str) -> int | None:
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx = {
        "medicine": col("Medicine name"),
        "substance": col("Active substance"),
        "date": col("Date of designation / refusal"),
        "use": col("Intended use"),
        "number": col("EU designation number"),
        "status": col("Status"),
        "url": col("Orphan designation URL"),
    }

    started = False
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not started:
            if "Medicine name" in cells:
                started = True
            continue

        def get(key: str) -> str | None:
            i = idx[key]
            if i is None or i >= len(cells):
                return None
            return cells[i] or None

        number = get("number")
        if not number:
            continue

        yield StagedDrug(
            designation_number=number,
            medicine_name=get("medicine"),
            active_substance=get("substance"),
            intended_use=get("use"),
            status=get("status"),
            designation_date=get("date"),
            url=get("url"),
        )


# ------------------------------------------------------- emparejar con enfermedad

# Frases con las que la EMA envuelve el nombre de la enfermedad. Se recortan para
# quedarnos con la enfermedad en sí.
_PREFIXES = re.compile(
    r"^(treatment of|prevention of|treatment and prevention of|therapy of|"
    r"tratamiento de|prevención de)\s+",
    re.IGNORECASE,
)

_STOPWORDS = {
    "the", "of", "and", "in", "for", "with", "a", "an",
    "de", "la", "el", "los", "las", "y", "en", "con",
}


# Genitivo sajón. Va fuera ANTES de limpiar la puntuación, y no es un detalle
# cosmético: la EMA escribe "Wilson's disease" y Orphanet "Wilson disease". Si el
# apóstrofe se convierte en espacio quedan "wilson s disease" y "wilson disease", que
# no casan, y el fármaco se pierde. Quitando el posesivo, ambos dan "wilson disease".
_POSSESSIVE = re.compile(r"'s\b|'s\b|s'\b", re.IGNORECASE)


def normalize(text: str) -> str:
    """Normaliza para comparar: sin acentos, sin posesivos, sin puntuación ni vacías."""
    text = _PREFIXES.sub("", text.strip())
    text = _POSSESSIVE.sub("", text)
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    tokens = [t for t in text.split() if t and t not in _STOPWORDS]
    return " ".join(tokens)
